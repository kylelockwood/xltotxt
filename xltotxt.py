#! python3
# xltotxt copies and converts an Excel sheet to a .txt file

import openpyxl
import sys, os, re

USAGE = '\nUsage: xltotxt <source path\\file.extension> <sheet name> <target path\\file.txt>'
HELP = (USAGE + '\n\n'
        'Typing \'list\' or leaving an argument blank will list available actions.\n'
        'Leaving a directory blank assumes the current working directory.\n'
        'File or sheet names that contain spaces must be within quotes, i.e "book name.xlsx".\n'
        'File extension is required for reading file, but not for writing (.txt).\n'
        'Not all Excel file types are supported by openpyxl.'
        )
print('')

def ma_error(message):
# Missing argument error message
    print(f'Error: Missing argument {message}{USAGE}')
    return

def inv_error(choice):
# Invalid choice error message
    print(f'\'{choice}\' is not a valid option')
    return

def load_xl(sheet):
# Get values from excel sheet and return as a list of rows
    print(f'Loading data from {sheet}...')
    xlData = []
    for row in range(1, sheet.max_row + 1):
        xlRowData = ''
        for column in range(1, sheet.max_column + 1):
            xlValue = sheet.cell(row,column).value
            xlRowData = xlRowData + str(xlValue) + ' '                          # Cell data seperator is a space (' ') but could be changed to a comma for csv
        xlData.append(xlRowData)
    if len(xlData) == 1:                                                        # Blank sheets will contain one value of 'None'
        sys.exit(f'{sheet} contains no data')
    return xlData


# Insure proper number of arguments
if len(sys.argv) > 4:
    sys.exit(f'Error: Too many arguments.{USAGE}')


# Source file handler
try:
    sourceFile = str(sys.argv[1])
except:
    sourceFile = 'missing argument error'
    pass
if not re.match(r'^[a-zA-Z]:\\', sourceFile):                                   # Does fileName start with a drive letter?
    sourceFile = str(os.getcwd() + '\\' + sourceFile)                           # If not, use current directory instead
sourceName = os.path.basename(sourceFile)

try:
    wb = openpyxl.load_workbook(sourceFile)
except: 
    if sourceName.lower() == 'help':
        sys.exit(HELP)
    elif sourceName == 'missing argument error':
        ma_error('<source path\\file.extension>')
    elif not sourceName.lower() == 'list':                                      # 'list' returns all excel files in directory                        
        print(f'Error: Could not open file \'{sourceFile}\'')
        if not re.search('.xls',sourceFile):                                    # Check file type is excel
            sys.exit(f'Error: Source file not a valid Excel type.')
    sourcePath = os.path.dirname(os.path.abspath(sourceFile))
    xlfiles = []
    for files in os.listdir(sourcePath):                                        # list .xls files in directory
        if re.search('.xls', files):
            xlfiles.append(files)
    if len(xlfiles) == 0:
        sys.exit(f'No Excel files found in {sourcePath}')
    print(f'Excel files found in {sourcePath}:')
    for files in xlfiles:
        print(f'  {os.path.basename(files)}')
    sys.exit()


# Sheet handler
try:
    sheetName = str(sys.argv[2])
except:
    sheetName = 'missing sheet error'
    pass
try:
    sheet = wb[sheetName]
except:
    if sheetName == 'missing sheet error':
        ma_error('<sheet name>')
    elif not sheetName == 'list':                                               # 'list' returns names of all sheets in workbook                                     
        print(f'Sheet \'{sheetName}\' not found in \'{sourceName}\'')
    print(f'Sheet names in \'{sourceName}\':')
    for sheets in wb.sheetnames:
        print(f'  {sheets}')
    sys.exit()


# Target file handler
try:
    targetFile = str(sys.argv[3])
except:
    sys.exit(ma_error('<target path\\file.txt>'))
if not re.match(r'^[a-zA-Z]:\\', targetFile):                                   # Does fileName start with a drive letter?
    targetFile = str(os.getcwd() + '\\' + targetFile)                           # If not, use current directory instead
if not targetFile.endswith('.txt'):
    targetFile = targetFile + '.txt'                                            # If not txt, then make it txt

targetName = os.path.basename(targetFile)
status = 'Creating'                                                             # Default print status for when the file is being created


# Target file exists
choice = 0
if os.path.exists(targetFile):
    while True:
        print(f'\'{targetName}\' already exists.')
        while True:
            choice = input(f'Type \'1\' to append\n'
                            'Type \'2\' to overwrite\n'
                            'Type \'3\' to quit\n')
            try:
                choice = int(choice)
                if choice in range(1,4):
                    break
            except:
                pass
            inv_error(choice)

        if choice == 1: # Append
            status = 'Appending'
            break
        
        if choice == 2: # Overwrite
            while True:
                yn = str(input(f'Overwrite \'{targetName}\', are you sure? (Y/N) '))
                if yn.lower() == 'y':
                    status = 'Overwriting'
                    open(targetFile, 'w').close                                 # Erase targetFile
                    break              
                elif yn.lower() == 'n':
                    print('Overwrite action cancelled\n')
                    break                                                       # Goto 'Already exists' choices again
                else:
                    inv_error(yn)          
            if yn == 'y':                                                       # Overwrite by creating a new file (below)
                break

        if choice == 3: # Quit
            sys.exit()


# Write the text file
print('')
xlData = load_xl(sheet)
with open(targetFile, 'a') as txtFile:
    if choice == 1:
        txtFile.write('\n')
    print(f'{status} \'{targetName}\'...')
    for row in xlData:
        txtFile.write(str(row + '\n'))
    txtFile.close()
print('Done.')


# View the new txt file in the terminal?
while True:
    yn = str(input(f'\nDisplay \'{targetName}\' contents? (Y/N) '))
    if yn.lower() == 'y':
        print(f'\n-- {targetName.upper()} --\n')
        with open(targetFile, 'r') as f:
            lines = f.readlines()
            for line in lines:
                line = line.rstrip()                                            # Remove blank lines
                if line:
                    print(line)
            f.close()
        sys.exit('\n-- END OF FILE --')
    elif not yn.lower() == 'n':
        inv_error(yn)
    else:
        break